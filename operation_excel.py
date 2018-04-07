from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter

class OperationExcel(Workbook):
    def __init__(self):
        self.Sheet=self.create_sheet(0)

    def add_data(self,DataList):
        '''循环插入数据'''
        for Data in DataList:
            # for Col in range(1,Data):
            #     ColNum=get_column_letter(Col)
            #     self.Sheet.cell()
            self.Sheet.append(Data)