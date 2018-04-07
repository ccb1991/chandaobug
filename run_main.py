import analyzehtml

from openpyxl import load_workbook
import os

parentdir=os.path.dirname(__file__)
Html_dir=parentdir+r"/html"
Excel_dir=parentdir+r"/excel/PDS-BE待解决bug处理-陈辰柄180328.xlsx"

WookBook=load_workbook(Excel_dir)
WookSheet=WookBook['sheet']
for Root, Dirs, Files in os.walk(Html_dir):
    for File in Files:
        SheetName=File[:-4]
        # Sheet=WookBook.create_sheet(SheetName,0)
        Sheet=WookBook.copy_worksheet(WookSheet)
        Sheet.title=SheetName
        Sheet.column_dimensions['C'].width=90
        DataList=analyzehtml.AnalyzeHtml(Html_dir).analyze_html(File)
        for Data in DataList:
            Sheet.append(Data)
WookBook.save(Excel_dir+".xlsx")


